from openpyxl import Workbook
from openpyxl import load_workbook
# wb = Workbook()
#
# ws = wb.active
# ws['A1'] = 34
# wb.save("sample.xlsx")

wb = load_workbook('sample.xlsx')
ws = wb.active
val = ws['A1'].value
print(val)